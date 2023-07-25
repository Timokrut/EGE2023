import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
from openpyxl.styles import PatternFill

with pd.ExcelWriter('data.xlsx') as writer:
    
    # данные 
    url_additions = ['1_38_1_1_1_f', '1_17_1_1_1_f', '1_18_1_1_1_f', '1_32_1_1_1_f', '1_121_1_1_1_f']
    id_additions = ['52', '7', '15', '38', '169']
    naming = ['11.03.02', '09.03.01', '09.03.02', '10.03.01', '27.03.03']
    free_places_for_painting = [88, 78, 54, 36, 16]
    
    for k in range(5):

        # база
        url = 'https://priem.guap.ru/lists/' + url_additions[k]
        page = requests.get(url)
        soup = bs(page.text, 'html.parser')

        # создаем список с загаловками
        table1 = soup.find('table', id = 'tablestat' + id_additions[k])
        headers = []
        for i in table1.find_all('th'):
            title = i.text
            headers.append(title)

        # меняю названия заголовков
        headers[1], headers[3], headers[6] = 'СНИЛС', 'БАЛЛЫ', 'ОРИГИНАЛ'

        # создаем верхушку
        mydata = pd.DataFrame(columns=headers)

        # заполняем данными
        for j in table1.find_all('tr')[1:]:
            row_data = j.find_all('td')
            row = [i.text for i in row_data]
            lenght = len(mydata)
            mydata.loc[lenght] = row

        # удаляю лишние строки
        mydata.drop('п/п', inplace=True, axis=1)
        mydata.drop('Сумма баллов ВИ', inplace=True, axis=1)
        mydata.drop('Баллы за достижения', inplace=True, axis=1)

        # сортировка баллов по убыванию
        mydata.sort_values(by='БАЛЛЫ', ascending=False, inplace=True)
        
        # сохраняем на листе
        mydata.to_excel(writer, sheet_name = naming[k], index=False)

        # заливка бюджетных мест
        if free_places_for_painting[k] <= len(mydata):
            sheet = writer.sheets[naming[k]]
            for row_idx in range(2, free_places_for_painting[k] + 2):  # Учитываем заголовки, начинаем с 2 строки
                for col_idx in range(1, 5):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color="caff9b", end_color="caff9b", fill_type="solid")




print('success')
